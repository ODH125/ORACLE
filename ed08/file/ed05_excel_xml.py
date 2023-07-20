'''

'''

from openpyxl import load_workbook
def main():
    wb = load_workbook('KOBIS_주간_주말_박스오피스_2023-07-19.xlsx')
    ws = wb.active # 첫번째 시트 활성

    for row in ws.iter_rows(min_row=9,values_only=True):
        # print('row',row)
        순위 = row[0]
        영화명 = row[1]
        점유율 = round(row[4]*100,2)
        관객수 = row[8]
        print(f'순위:{순위}, 영화명:{영화명}, 점유율:{점유율}, 관객수:{관객수}')
main()
